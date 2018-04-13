"""
This program is a program I made for my mother. I used Python Tkinter to make the GUI and the commands. And I used OpenPyXL 
to transfer the items to be delivered from the app into Microsoft Excel so that it can be printed or edited in a excel sheet.

@author: Potterishpotter 04/06/2018 3:00AM
""'
from tkinter import *
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Border, Alignment, Side, Font
# imported datetime so that I can use the date today and set the file name to the date today so that the confussions can be avoided
import datetime

try: # tries to read the file medicineList where the medicine list will be stored 
    file = open('medicineList.txt', 'r')
    file.close()
except: # and if the file is not readable which means it doesn't exist, it will create that file.
    file = open('medicineList.txt', 'w+')
    file.close()

# default border which makes all of the border of 1 cell have a thin border in the excel file
defaultBorder = Border(left=Side(style='thin'), right = Side(style='thin'), top = Side(style='thin'), bottom = Side(style='thin'))

# center alignment basically just aligns the text in the cells in the middle, both horizontally and vertically
centerAlignment = Alignment(horizontal="center", vertical="center")

root = Tk()
root.title('Delivery Receipt')
root.geometry('635x510')
root.resizable(False, False)

#=============================================FUNCTIONS===================================================
def addExcel(*args): # the function that adds all the data from the DeliverTo listbox into the excel file
    global centerAlignment
    global defaultBorder

    wb = Workbook()
    ws = wb.active

    ws.page_margins.left = 0.1
    ws.page_margins.top = 0.1

    ws.column_dimensions['A'].width = 3.78
    ws.column_dimensions['B'].width = 6.78
    ws.column_dimensions['C'].width = 8.78
    ws.column_dimensions['D'].width = 8.78
    ws.column_dimensions['E'].width = 8.78
    ws.column_dimensions['F'].width = 6.78

    ws.row_dimensions[4].height = 20

    ws.merge_cells('A1:F2')
    ws.merge_cells('A3:B3')
    ws.merge_cells('E3:F3')
    ws.merge_cells('C4:D4')
    ws['A1'] = 'Delivery Receipt'
    ws['A3'] = 'Deliver to'
    ws['B3'] = str(deli.get())
    ws['D3'] = 'Date'
    ws['E3'] = str(d.get())

    ws['A4'] = 'Qty.'
    ws['B4'] = 'Unit'
    ws['C4'] = 'Description'
    ws['E4'] = 'U/Price'
    ws['F4'] = 'Amount'

    styleDesign(ws, 'A4', defaultBorder, centerAlignment)
    styleDesign(ws, 'B4', defaultBorder, centerAlignment)
    styleDesign(ws, 'C4', defaultBorder, centerAlignment)
    styleDesign(ws, 'D4', defaultBorder, centerAlignment)
    styleDesign(ws, 'E4', defaultBorder, centerAlignment)
    styleDesign(ws, 'F4', defaultBorder, centerAlignment)

    ws['A1'].alignment = centerAlignment
    ws['A1'].font = Font(size=22)

    # ws['A1'].border = defaultBorder
    # ws['B1'].border = defaultBorder
    # ws['C1'].border = defaultBorder
    # ws['D1'].border = defaultBorder
    # ws['E1'].border = defaultBorder
    # ws['F1'].border = defaultBorder
    # ws['A2'].border = defaultBorder
    # ws['B2'].border = defaultBorder
    # ws['C2'].border = defaultBorder
    # ws['D2'].border = defaultBorder
    # ws['E2'].border = defaultBorder
    # ws['F2'].border = defaultBorder
    # ws['A3'].border = defaultBorder
    # ws['B3'].border = defaultBorder
    # ws['C3'].border = defaultBorder
    # ws['D3'].border = defaultBorder
    # ws['E3'].border = defaultBorder
    # ws['F3'].border = defaultBorder



    count = 5

    for x in enumerate(toDeliver.get(0, END)):
        data = x[1].split(' ')
        q = data[0]
        u = data[1]
        i = data[2]


        if count <= 29:
            ws.merge_cells('C' + str(count) + ':D' + str(count))

            ws['A' + str(count)] = q
            ws['B' + str(count)] = u
            ws['C' + str(count)] = i

            styleDesign(ws, 'A' + str(count), defaultBorder, centerAlignment)
            styleDesign(ws, 'B' + str(count), defaultBorder, centerAlignment)
            ws['C' + str(count)].border = defaultBorder
            ws['D' + str(count)].border = defaultBorder
            styleDesign(ws, 'E' + str(count), defaultBorder, centerAlignment)
            styleDesign(ws, 'F' + str(count), defaultBorder, centerAlignment)
        elif count % 6 == 0:
            count += 2

        count += 1


    ws.merge_cells('A' + str(count) + ':B' + str(count))
    ws['A' + str(count)] = 'Prepared by: '

    ws.merge_cells('D' + str(count) + ':E' + str(count))
    ws['D' + str(count)] = 'Received by: '

    wb.save('DeliveryReceipt' + str(datetime.date.today()) + '.xlsx')



def styleDesign(ws, cell, border, alignment): # the style of each cell
    ws[cell].border = border
    ws[cell].alignment = alignment

def addList(*args): # add a new item into the medicineList
    global mm

    if m.get() in medicines.get(0, END):
        alreadyInList = Tk()
        alreadyInList.title('Alert!')
        aFrame = ttk.Frame(alreadyInList, padding='10 10 10 10')
        aFrame.grid(column=0, row=0)
        ttk.Label(aFrame, text="'" + m.get() + "'" + ' is already inside your list of to be medicines items.', font=('', 12)).grid(column=0, row=1)
        pass
    else:
        add.write(m.get() + ' ')
        listx = list(mm)
        listx.append(m.get())
        mm = tuple(listx)
        medicines.insert(END, m.get())
        print('Added ' + str(m.get()) + ' in medicines')


    addMed.delete(0, 'end')


def addSelected(*args): # add an item from the medicineList into the deliverTo list
    idxs = medicines.curselection()
    if mm[idxs[0]] in selected.get():
        alreadyInList = Tk()
        alreadyInList.title('Alert!')
        aFrame = ttk.Frame(alreadyInList, padding='10 10 10 10')
        aFrame.grid(column=0, row=0)
        ttk.Label(aFrame, text="'" + mm[idxs[0]] + "'" + ' is already inside your list of to be delivered items.', font=('', 12)).grid(column=0, row=1)
    else:
        toDeliver.insert(END, str(qty.get()) + ' ' + str(uni.get()) + ' ' + str(mm[idxs[0]]))
        print('Added ' + str(qty.get()) + ' ' + str(uni.get()) + ' ' + str(mm[idxs[0]]) + ' to selected items')

def reMed(*args): # removes a medicine from the medicineList listbox
    global mm
    idxs = medicines.curselection()
    listx = list(mm)
    del listx[idxs[0]]
    mm = tuple(listx)
    medicines.delete(idxs[0])
    print('Removed index ' + str(idxs[0]) + ' from medicines')


def reSel(*args): # removes a selected item from the deliverTo listbox
    idxs = toDeliver.curselection()
    toDeliver.delete(idxs[0])
    print('Removed index ' + str(idxs[0]) + ' from selected items')

print('Log: ')

#=============================================FUNCTIONS===================================================
medi = open('medicineList.txt', 'r') # opens medicineList as readable
listx = list(medi.read()) # sets the data from medicineList into a list
allString = ''.join(listx) # then joins them because when you make it a list, it lists it letter by letter, like ['a', 'b', 'c']
listx = allString.split(' ') # then splits them because there are whitespaces in every word.
listx.pop() #without this there would be an extra space in the choices
listx.sort() # sorts the list
mm = tuple(listx) # then finally converts the list into a tuple
medi.close()

med = StringVar(value=mm)

frame = ttk.Frame(root, padding='5 5 5 5')
medicines = Listbox(frame, listvariable=med, width=100, height=10)
msc = ttk.Scrollbar(frame, orient=VERTICAL, command=medicines.yview)
medicines['yscrollcommand'] = msc.set


#=============================info frame============================
infoFrame = ttk.Frame(frame, padding='3 3 3 3')

delivert = ('GS Unifarm', 'C&C Farm')
deli = StringVar()
d = StringVar() # date

lblDeliverTo = ttk.Label(infoFrame, text='Deliver to: ')
deliverTo = ttk.Combobox(infoFrame, value=delivert, textvariable=deli)
lblDate = ttk.Label(infoFrame, text='Date: ')
date = ttk.Entry(infoFrame, textvariable=d)

m = StringVar() # medicine

addMed = ttk.Entry(frame, textvariable=m)
addToList = ttk.Button(frame, text='Add to List', command=addList)
removeMed = ttk.Button(frame, text='Remove from Medicine', command=reMed)

#=============================add frame==============================

addFrame = ttk.Frame(frame, padding='3 3 3 3')

un = ('bottle/s', 'box', 'boxes')
uni = StringVar()
qty = StringVar()

lblQuantity = ttk.Label(addFrame, text='Qty.')
quantity = Spinbox(addFrame, from_=1, to=100, textvariable=qty)
lblUnit = ttk.Label(addFrame, text='Unit')
unit = ttk.Combobox(addFrame, values=un, textvariable=uni)

addToDeliver = ttk.Button(addFrame, text='Add to Deliver', command=addSelected)

selected= StringVar()

toDeliver = Listbox(frame, listvariable=selected)
csc = ttk.Scrollbar(frame, orient=VERTICAL, command=toDeliver.yview)
toDeliver['yscrollcommand'] = csc.set

removeSelected = ttk.Button(frame, text='Remove from Selected', command=reSel)
printt = ttk.Button(frame, text='Print', command=addExcel)

frame.grid(column=0, row=0)

ttk.Label(frame, text='Medicine: ', font=('', 20)).grid(column=0, row=0, rowspan=2, sticky=(N, W))

infoFrame.grid(row=0, column=3, rowspan=2, sticky=E)
lblDeliverTo.grid(column=0, row=0, sticky=E, pady=0.5)
deliverTo.grid(column=1, row=0, sticky=(E, W), pady=0.5)
lblDate.grid(column=0, row=1, sticky=E, pady=0.5)
date.grid(column=1, row=1, sticky=(E, W), pady=0.5)

medicines.grid(column=0, row=2, columnspan=4, rowspan=2)
msc.grid(column=4, row=2, rowspan=2, sticky=(N, S))

addFrame.grid(column=3, row=5, sticky=(S, E))
lblQuantity.grid(column=0, row=0)
quantity.grid(column=0, row=1)
lblUnit.grid(column=1, row=0)
unit.grid(column=1, row=1)
addToDeliver.grid(row=2, column=0, columnspan=2, sticky=(E, W))

addMed.grid(column=0, columnspan=4, row=4, sticky=(E, W))
addToList.grid(column=2, row=4, sticky=(E, W))
removeMed.grid(column=3, row=4, sticky=(E, W))

ttk.Label(frame, text='Selected: ', font=('', 20)).grid(column=0, row=5, sticky=(S, W), pady=3)
toDeliver.grid(column=0, row=6, columnspan=4, rowspan=2, sticky=(E,W))
csc.grid(column=4, row=6, rowspan=2, sticky=(N, S))
removeSelected.grid(column=2, row=8, sticky=(S, W, E))
printt.grid(column=3, row=8, sticky=(S, W, E))



add = open('medicineList.txt', 'a')
root.mainloop()
add.close()
